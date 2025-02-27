# -*- coding=utf-8 -*-
import os, re, time
import openpyxl
from openpyxl.styles import PatternFill  # , Border, Side, Alignment, Protection, Font,fills,colors


# from PyQt5.QtWidgets import QFileDialog

def getKeyWordFromSettingFile(config_details):
    TER_keyword = {'Platform': '', 'Item': '', 'SiteNum': '', 'PatName': '', 'PlotStart': '', 'PlotEnd': '',
                   'RowOffset': '', 'PassSymbol': '', 'FailSymbol': ''}
    # keyword = []
    #
    # json_file = SettingFilePath
    # if json_file:
    #     # Load config values
    #     with open(json_file) as config_file:
    #         config_details = json.load(config_file)

    TER_keyword['Platform'] = 'TER'
    TER_keyword['Item'] = config_details["keyword_item"]
    TER_keyword['SiteNum'] = config_details["keyword_site"]
    # TER_keyword['PatName'] = keyword[3]
    TER_keyword['PlotStart'] = config_details["keyword_start"]
    TER_keyword['PlotEnd'] = config_details["keyword_end"]
    TER_keyword['RowOffset'] = '1'
    TER_keyword['PassSymbol'] = config_details["keyword_pass"]  # keyword[7]
    TER_keyword['FailSymbol'] = config_details["keyword_fail"]  # keyword[8]
    # print(TER_keyword)

    return TER_keyword


def getDatalogInfo(TER_keyword, file_paths:str, site_lbl:str):
    totalsiteCnt = 0
    str_sites = []
    site_lbl_list = site_lbl.split(';')

    # get process site label
    select_sites = ''
    TER_files = file_paths.split(';')
    for i in range(len(TER_files)):
        # select_sites = str(input(each_file + '  ' + 'select sites:  '))
        select_sites = site_lbl_list[i]
        if select_sites == "":
            for each_file in TER_files:
                select_sites = getAllSiteNums(each_file, TER_keyword)
        totalsiteCnt = totalsiteCnt + len((select_sites.split(',')))
        str_sites.append(select_sites)

    print("totalsiteCnt= " + str(totalsiteCnt))
    xls = openpyxl.Workbook()
    siteCnt = 0
    index = 0
    for each_file in TER_files:
        sites_TER = str_sites[index].split(',')
        index = index + 1
        for each_site in sites_TER:
            siteCnt = siteCnt + 1
            processLog(each_file, each_site, xls, siteCnt, totalsiteCnt, TER_keyword)

    time_flag = (time.strftime('%Y_%m_%d_%H_%M_%S', time.localtime(time.time())))
    report_file = os.getcwd() + '/shmplot_' + time_flag + '.xlsx'
    xls.save(report_file)
    return report_file


def processLog(each_file, each_site, xls, siteCnt, totalsiteCnt, dict_keyword):
    # time_flag = (time.strftime('%Y_%m_%d_%H_%M_%S', time.localtime(time.time())))
    flag = 0
    startPlot = 0
    site_flag = 0
    each_item_info = []
    each_plot = []
    file = open(each_file, 'r', encoding='utf-8')
    shtName = 'Sheet'

    iRow = 2
    interval_columns = 25
    iColumn = (siteCnt - 1) * interval_columns
    for line in file.readlines():
        if ('VBT error' in line):
            continue
        # if (dict_keyword['Item'] in line):
        if (re.match(dict_keyword['Item'], line) != None):
            each_item_info.append(line)
            flag = 1
        if (dict_keyword['SiteNum'] in line and not (
                (dict_keyword['SiteNum'] + str(each_site)) == line[:-1]) and flag == 1):
            each_item_info = []
            site_flag = 0

        if (dict_keyword['SiteNum'] + str(each_site)) == line[:-1] and flag == 1:
            each_item_info.append(line)
            site_flag = 1
            continue
        if line and flag == 1 and site_flag == 1 and startPlot == 0: #dict_keyword['PatName'] in line and flag == 1 and site_flag == 1:
            each_item_info.append(line)

        if (line.strip().startswith(dict_keyword['PlotStart']) and flag == 1 and site_flag == 1):
            # (dict_keyword['PlotStart'] in line.strip()  and flag==1 and site_flag==1) :
            startPlot = 1
        if startPlot == 1 and flag == 1 and site_flag == 1:
            if line != '\n':
                each_plot.append(line)
        if (dict_keyword['PlotEnd'] in line and flag == 1 and startPlot == 1 and site_flag == 1):
            xls[shtName].cell(row=1, column=(siteCnt - 1) * interval_columns + 1, value=each_file)
            for tmpstr in each_item_info:
                iColumn = (siteCnt - 1) * interval_columns
                xls[shtName].cell(row=iRow + 1, column=iColumn + 1, value=tmpstr)
                xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + 1, value=tmpstr)
                iRow = iRow + 1
                pass
            for each_plot_line in each_plot:
                iColumn = (siteCnt - 1) * interval_columns
                # tmpstr = each_plot_line.split('\t')
                tmpstr = each_plot_line.split()
                if not (dict_keyword['PlotEnd'] in each_plot_line):
                    # Change Pass/Fail symbol to 'P' and '.'
                    for i in range(1, len(tmpstr)):
                        tmp = tmpstr[i]
                        text_pass = dict_keyword['PassSymbol']
                        text_fail = dict_keyword['FailSymbol']
                        if re.match(text_pass, tmp) != None:
                            tmpstr[i] = 'P'
                        if re.match(text_fail, tmp) != None:
                            tmpstr[i] = '.'

                for x in tmpstr:
                    xls[shtName].cell(row=iRow + 1, column=iColumn + 1, value=x)
                    if x == 'P':
                        fill = PatternFill(start_color='00EE00', end_color='00EE00', fill_type="solid", )
                        xls[shtName].cell(row=iRow + 1, column=iColumn + 1).fill = fill
                    elif x == '.':
                        fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type="solid", )
                        xls[shtName].cell(row=iRow + 1, column=iColumn + 1).fill = fill
                    if str(xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + (
                            iColumn - (siteCnt - 1) * interval_columns) + 1).value) != 'None':
                        y = str(xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + (
                                    iColumn - (siteCnt - 1) * interval_columns) + 1).value) + x
                        xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + (
                                    iColumn - (siteCnt - 1) * interval_columns) + 1, value=y)
                        if 'P' in y and '.' in y:
                            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid", )
                            xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + (
                                        iColumn - (siteCnt - 1) * interval_columns) + 1).fill = fill
                    else:
                        xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + (
                                    iColumn - (siteCnt - 1) * interval_columns) + 1, value=x)
                    iColumn = iColumn + 1
                iRow = iRow + 1
                pass

            startPlot = 0
            flag = 0
            site_flag = 0
            each_item_info = []
            each_plot = []

            iRow = iRow + int(dict_keyword['RowOffset'])
            # if dict_keyword['Platform'] == 'TER':
            #     iRow = iRow + int(TER_keyword['RowOffset'])
            # elif dict_keyword['Platform'] == 'ADV':
            #     iRow = iRow + int(ADV_keyword['RowOffset'])

    # xls.save(os.getcwd() + '/shmplot_'+time_flag+'.xlsx')


def getAllSiteNums(each_file, TER_keyword):
    # global TER_keyword
    # global ADV_keyword
    site_info = []
    str_site = ''
    unique_site_info = []
    file = open(each_file, 'r', encoding='utf-8')
    for each_line in file.readlines():
        if TER_keyword['SiteNum'] != "" and TER_keyword['SiteNum'] in each_line:
            site_info.append(int(each_line[len(TER_keyword['SiteNum']):len(each_line)].strip()))

    unique_site_info = list(set(site_info))
    # unique_site_info=list.sort(unique_site_info)

    for x in unique_site_info:
        if str_site == '':
            str_site = str(x)
        else:
            str_site = str_site + ',' + str(x)

    return str_site


if __name__ == '__main__':
    TER_keyword, ADV_keyword = getKeyWordFromSettingFile()
    getDatalogInfo(TER_keyword, ADV_keyword)




